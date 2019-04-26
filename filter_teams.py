from openpyxl import load_workbook, Workbook


def filter_teams():

    wb = load_workbook(filename="ncaa_basketball.xlsx")

    print(wb.sheetnames)

    year = 2019

    for sheet in wb.worksheets:
        print("Filtering Sheet {}".format(year))
        filter_sheet(sheet)
        year -= 1

    wb.save(filename="ncaa_basketball_filtered.xlsx")


def create_dataset():
    wb = load_workbook(filename="ncaa_basketball_filtered.xlsx")

    year = 2019

    f = open("ncaa_tournament.data.csv", "w+")

    for sheet in wb.worksheets:
        print("Extracting Sheet {}".format(year))
        year -= 1

        sheet.delete_rows(1, 1)
        sheet.delete_cols(1, 1)

        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column):
                f.write(str(sheet.cell(i, j).value) + ",")
            f.write(str(sheet.cell(i, sheet.max_column).value) + '\n')

    f.close()


def filter_sheet(sheet):

    # Start searching at first team row
    row_index = 4
    offset = "B" + str(row_index)
    name = sheet[offset].value

    while name is not None:
        # print(name)
        if "NCAA" not in name:
            sheet.delete_rows(row_index)
            name = sheet[offset].value
        else:
            row_index += 1
            offset = "B" + str(row_index)
            name = sheet[offset].value

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            val = sheet.cell(i, j).value
            try:
                val = float(val)
            except ValueError:
                pass
            except TypeError:
                break

            sheet.cell(i, j).value = val

    sheet.delete_cols(9, 9)
    sheet.delete_cols(7, 1)
    sheet.delete_cols(3, 3)
    sheet.delete_cols(1, 1)


if __name__ == '__main__':
    create_dataset()
